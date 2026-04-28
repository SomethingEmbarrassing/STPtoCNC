from pathlib import Path

from openpyxl import load_workbook

from stptocnc.workflows import finalize_nest_run
from stptocnc.nesting import pack_instances_first_fit, move_instance_between_nests
from stptocnc.models.nesting import PartInstance, EndCondition
from stptocnc.config import NestingDefaults, ProfileFamily


def test_finalize_nest_run_creates_cutlist(tmp_path: Path) -> None:
    cutlist = tmp_path / "final-cutlist.xlsx"
    cnc_dir = tmp_path / "cnc"

    result = finalize_nest_run(
        nc1_files=["docs/pp1016.nc1", "docs/as1007.nc1"],
        cutlist_output=cutlist,
        cnc_output_dir=cnc_dir,
    )

    assert result["status"] == "ok"
    generated_cutlist = Path(result["cutlist"])
    assert generated_cutlist.exists()
    assert generated_cutlist.name.startswith("final-cutlist_")
    assert generated_cutlist.suffix == ".xlsx"
    assert len(result["cnc_files"]) >= 1
    first_cnc = Path(result["cnc_files"][0])
    cnc_text = first_cnc.read_text(encoding="utf-8")
    assert "(PLACEHOLDER NESTED CNC)" not in cnc_text
    assert "(POST EMI 2400 PROMPTS ROP V1.4)" in cnc_text
    assert "(PIECE " in cnc_text

    wb = load_workbook(generated_cutlist)
    assert wb.sheetnames == ["CutList"]


def test_finalize_uses_prepared_nests_for_manual_reassignment(tmp_path: Path) -> None:
    parts = [
        PartInstance("A", "A", 144.0, ProfileFamily.PIPE, EndCondition.FLAT, EndCondition.FLAT),
        PartInstance("B", "B", 144.0, ProfileFamily.PIPE, EndCondition.FLAT, EndCondition.FLAT),
        PartInstance("C", "C", 36.0, ProfileFamily.PIPE, EndCondition.FLAT, EndCondition.FLAT),
    ]
    defaults = NestingDefaults()
    original = pack_instances_first_fit(parts, defaults).nests
    moved = move_instance_between_nests(original, "C", "nest-2", defaults)
    result = finalize_nest_run(
        nc1_files=[],
        cutlist_output=tmp_path / "manual-cutlist.xlsx",
        cnc_output_dir=tmp_path / "cnc",
        prepared_nests=moved,
    )
    assert result["status"] == "ok"
    nest2 = Path(tmp_path / "cnc" / "nest-2.cnc").read_text(encoding="utf-8")
    assert "(PIECE 2: C START=144.000 LEN=36.000)" in nest2
