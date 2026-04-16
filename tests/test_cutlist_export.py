from pathlib import Path

from openpyxl import load_workbook

from stptocnc.models.nesting import LinearNest, NestPlacement
from stptocnc.config import ProfileFamily
from stptocnc.reports import normalize_material_shape, write_cutlist_workbook
from stptocnc.reports.cutlist_xlsx import format_inches_fraction


def _build_sample_nests() -> list[LinearNest]:
    nest1 = LinearNest(nest_id="nest-1", profile_family=ProfileFamily.PIPE, stock_length_in=252.0)
    nest1.placements = [
        NestPlacement(
            instance_id="A#1",
            part_mark="A",
            offset_in=0.0,
            length_in=20.0,
            transition_trim_before_in=0.0,
            profile_designation="PIPE1-1/2SCH40",
            source_file="docs/pp1016.nc1",
        ),
        NestPlacement(
            instance_id="B#1",
            part_mark="B",
            offset_in=20.25,
            length_in=10.0,
            transition_trim_before_in=0.25,
            transition_reason="previous_end_not_flat_compatible",
            profile_designation="PIPE1-1/2SCH40",
            source_file="docs/pp1016.nc1",
        ),
    ]

    nest2 = LinearNest(nest_id="nest-2", profile_family=ProfileFamily.HSS, stock_length_in=240.0)
    nest2.placements = [
        NestPlacement(
            instance_id="C#1",
            part_mark="C",
            offset_in=0.0,
            length_in=30.0,
            transition_trim_before_in=0.0,
            profile_designation="HSS8X4X1/4",
            source_file="docs/h1001.nc1",
        )
    ]
    return [nest1, nest2]


def test_cutlist_workbook_created_and_layout(tmp_path: Path) -> None:
    output = tmp_path / "cutlist.xlsx"
    write_cutlist_workbook(_build_sample_nests(), output)

    assert output.exists()

    wb = load_workbook(output)
    assert wb.sheetnames == ["CutList"]
    ws = wb["CutList"]

    values = [row for row in ws.iter_rows(values_only=True)]
    flat = [cell for row in values for cell in row if cell is not None]

    assert "Nest Cut List" in flat
    assert "Total Pieces" in flat
    assert "Total Nests" in flat

    # headers
    assert "Nest ID" in flat
    assert "Piece Mark" in flat
    assert "Material Shape" in flat
    assert "Piece Length" in flat
    assert "Start Offset" in flat
    assert "Drop" in flat
    assert "Trim Cut" in flat

    # grouped by nest with cut order preserved
    nest_rows = [row for row in values if row and isinstance(row[0], str) and row[0].startswith("Nest: ")]
    assert nest_rows[0][0] == "Nest: nest-1.cnc"
    assert nest_rows[1][0] == "Nest: nest-2.cnc"

    detail_rows = [row for row in values if row and row[0] in {"nest-1.cnc", "nest-2.cnc"}]
    assert detail_rows[0][1] == 1
    assert detail_rows[1][1] == 2


def test_material_normalization_rules() -> None:
    assert normalize_material_shape("HSS8X4X1/4") == "HSS8X4X1/4"
    assert normalize_material_shape("HSS 8 X 4 X 1/4") == "HSS8X4X1/4"
    assert normalize_material_shape("L2X2X1/4") == "L2X2X1/4"
    assert normalize_material_shape("ANGLE 3X3X1/4") == "L3X3X1/4"
    assert normalize_material_shape("PIPE1-1/2SCH40") == "PIPE 1-1/2 SCH 40"
    assert normalize_material_shape("pipe 1.5 sch 80") == "PIPE 1.5 SCH 80"


def test_fractional_inch_formatting_and_stock_summary_shape(tmp_path: Path) -> None:
    output = tmp_path / "cutlist.xlsx"
    write_cutlist_workbook(_build_sample_nests(), output)
    ws = load_workbook(output)["CutList"]

    # Raw stock summary uses operator-readable shape with dimensions.
    assert "PIPE 1-1/2 SCH 40 @ 252\"" in str(ws["B6"].value)
    assert "HSS8X4X1/4 @ 240\"" in str(ws["B6"].value)

    # Detail rows use nearest 1/16" string display.
    detail_rows = [row for row in ws.iter_rows(values_only=True) if row and row[0] == "nest-1.cnc"]
    assert detail_rows[0][4] == '20"'
    assert detail_rows[1][7] == '1/4"'


def test_format_inches_fraction_rounds_to_sixteenth() -> None:
    assert format_inches_fraction(1.124) == '1 1/8"'
    assert format_inches_fraction(1.13) == '1 1/8"'
    assert format_inches_fraction(0.24) == '1/4"'
